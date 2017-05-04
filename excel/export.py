from collections import OrderedDict

import datetime

from openpyxl import Workbook


class ExcelExport:
    def __init__(self):
        self.work_book = Workbook()

        self.row_items = (
            ('date',), ('name',), ('price',), ('count',),
            ('category', 'name'))

    def _get_rows(self, query):
        rows = []
        for item_list in enumerate(query):
            value = []
            for row_item in enumerate(self.row_items):
                last_value = getattr(item_list[1], row_item[1][0])
                for attr in row_item[1][1:]:
                    last_value = getattr(last_value, attr)
                value.append(last_value)
            rows.append(value)
        return rows

    def create_file(self, file, query, sheet_name):
        items = self._get_rows(query)
        work_list = self.work_book.create_sheet(sheet_name)
        for item in enumerate(items):
            for value in enumerate(item[1]):
                col_item = value[1]
                if isinstance(col_item, datetime.date):
                    if item[0] > 0 and items[1][value[0]] == col_item:
                        work_list.merge_cells(start_row=item[0] + 1, start_column=value[0] + 1, end_row=item[0] - 1 + 1,
                                              end_column=value[0] + 1, )
                work_list.cell(row=item[0] + 1, column=value[0] + 1, value=col_item)
        self.work_book.save(self._get_file_name(file))

    def _get_file_name(self, file):
        if file.split('.')[-1] == 'xlsx':
            return file
        return file + '.xlsx'
